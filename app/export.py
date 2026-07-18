from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import PayrollRecord, User

BRAND = "e5006e"
DARK = "1a1a2e"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=DARK, size=14)
LABEL_FONT = Font(name="Calibri", bold=True, color=DARK, size=11)
VALUE_FONT = Font(name="Calibri", color="000000", size=11)
TOTAL_FONT = Font(name="Calibri", bold=True, color=BRAND, size=12)
HIGHLIGHT_FILL = PatternFill("solid", fgColor="fde8f2")
THIN_BORDER = Border(
    left=Side(style="thin", color="d0d0d8"),
    right=Side(style="thin", color="d0d0d8"),
    top=Side(style="thin", color="d0d0d8"),
    bottom=Side(style="thin", color="d0d0d8"),
)


def _money(v: float) -> float:
    return round(float(v or 0), 2)


def generate_payroll_xlsx(record: PayrollRecord, user: User) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт ЗП"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24

    ws.merge_cells("A1:B1")
    ws["A1"] = "Бит.Serves — Расчёт заработной платы"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    dept_name = user.department.name if user.department else "—"
    pos_name = user.position.name if user.position else "—"
    grade_name = record.grade.name if record.grade else record.grade_id

    info_rows = [
        ("Получатель", user.full_name or "—"),
        ("Отдел", dept_name),
        ("Должность", pos_name),
        ("Грейд", grade_name),
        ("Период", record.period),
        ("Дата расчёта", record.created_at.strftime("%d.%m.%Y %H:%M") if record.created_at else "—"),
    ]
    row = 2
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2, value=str(value)).font = VALUE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Параметры расчёта").font = Font(bold=True, size=12, color=BRAND)
    row += 1

    params = [
        ("Отработано дней", record.worked_days, "дн."),
        ("Рабочих дней в месяце", record.working_days, "дн."),
        ("Маржа с услуг", _money(record.service_margin), "₽"),
        ("Маржа с товара", _money(record.goods_margin), "₽"),
        ("Процент премии", float(record.bonus_percent), "%"),
        ("Коэффициент услуг", float(record.service_factor), ""),
        ("НДФЛ", float(record.tax_rate), "%"),
    ]
    for label, value, unit in params:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right")
        if unit == "₽":
            ws.cell(row=row, column=2).number_format = "#,##0.00"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Результат расчёта").font = Font(bold=True, size=12, color=BRAND)
    row += 1

    headers = ["Показатель", "Сумма, ₽"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    results = [
        ("Начислено по окладу", _money(record.accrued_base)),
        ("Премия за услуги", _money(record.services_bonus)),
        ("Премия за товар", _money(record.goods_bonus)),
        ("Премия итого", _money(record.bonus_total)),
        ("Начислено всего (gross)", _money(record.gross_pay)),
        ("НДФЛ", _money(record.tax_amount)),
    ]
    for label, value in results:
        ws.cell(row=row, column=1, value=label).font = VALUE_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = "#,##0.00"
        row += 1

    ws.cell(row=row, column=1, value="К выплате (net)").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = HIGHLIGHT_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    cell = ws.cell(row=row, column=2, value=_money(record.net_pay))
    cell.font = TOTAL_FONT
    cell.fill = HIGHLIGHT_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = "#,##0.00"
    ws.row_dimensions[row].height = 24

    last_col = get_column_letter(2)
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()